from openpyxl import Workbook
from pirc522 import RFID
import RPi.GPIO as GPIO    # Import Raspberry Pi GPIO library
from time import sleep     # Import the sleep function from the time module
import sys
import os
import time
import mysql.connector
import MySQLdb
from mysql.connector import Error
from mysql.connector import errorcode
from datetime import datetime

sys.path.insert(0, '/home/pi/PointRFID/Script/I2C-LCD/')
import lcddriver


lcd = lcddriver.lcd()

GPIO.setwarnings(False)    # Ignore warning for now


# Connect to DB -----------------------------------------------------------
db = MySQLdb.connect(host="localhost", user="root",passwd="MELEC", db="PointRFID")
cur = db.cursor()


DATE = time.strftime("%Y-%m-%d %H:%M:%S")

rdr = RFID()


out = open('/home/pi/PointRFID/Log/lecture.log','w')

sys.stdout = out
sys.stderr = out


while True:

	lcd.lcd_clear()
	lcd.lcd_display_string(" Pret ", 1)

	# Detecter les tags
	rdr.wait_for_tag()
	(error, tag_type) = rdr.request()
	if not error:
		print("Tag detected")
		(error, uid) = rdr.anticoll()
		if not error:
			RFID_UID = (str(uid[0])+""+str(uid[1])+""+str(uid[2])+""+str(uid[3]))
			print ("\nUID de la carte  :  "),RFID_UID
			sys.stdout.flush()
			lcd.lcd_clear()
			lcd.lcd_display_string(" Carte  ", 1)
			lcd.lcd_display_string(" Detectee ", 2)
			sleep(1)
			try:
				print ("\nLecture de la base Worker...")
				sys.stdout.flush()
				lcd.lcd_clear()
				lcd.lcd_display_string(" Lecture  ", 1)
				lcd.lcd_display_string(" En Cours ", 2)
				sleep(1)

				# Execute the SQL command
				recuperation = "SELECT * FROM WORKER WHERE RFID_UID = (%s) AND STATUT = 'Eleve'"

				cur.execute(recuperation, [RFID_UID])

				rows = cur.fetchall()

				ID = rows[0][0]
				NOM = rows[0][1]
				PRENOM = rows[0][2]
				EMAIL = rows[0][3]
				GROUPE = rows[0][4]
				ENTREPRISE = rows[0][5]
				UID = rows[0][6]
				HEURE = time.strftime("%Y-%m-%d %H:%M")
				EXPORT = 'False'
				

				presence = 0

				presence = "SELECT * FROM LOG WHERE RFID_UID = (%s) AND EXPORT = 'False'"
				presence = cur.execute(presence, [UID])

				print presence

				if presence == 1:
				#Suppression de la ligne
					try:

						print ("\nPersonne deja presente ")
						sys.stdout.flush()
						lcd.lcd_clear()
						lcd.lcd_display_string(" Personne  ", 1)
						lcd.lcd_display_string(" Presente ", 2)
						sleep(1)

					except:
						print ("\nErreur ")
						sys.stdout.flush()

				else:
					#Insertion de l'eleve
					try:
				
						insertion = """ INSERT INTO `LOG`(`NOM`, `PRENOM`, `GROUPE`, `ENTREPRISE`, `RFID_UID`, `HEURE`, `EXPORT`) VALUES (%s,%s,%s,%s,%s,%s,%s)"""

						insertion_var = (NOM, PRENOM, GROUPE, ENTREPRISE, RFID_UID, HEURE, EXPORT)
						
						result = cur.execute(insertion, insertion_var)

						data = cur.fetchall()

						db.commit()

						lcd.lcd_clear()
						lcd.lcd_display_string(" Bonjour  ", 1)
						lcd.lcd_display_string(" ..... ", 2)
						sleep(1)


						print ("\nPersonne ajouter a la base")
						sys.stdout.flush()
						
					except mysql.connector.Error as error :
						db.rollback()

						print("\nImpossible d\'executer la commande vers MySQL {}".format(error))
						sys.stdout.flush()
			except:
				try:
					print ("\nLecture de la base Worker...")
					sys.stdout.flush()
					# Execute the SQL command
					test_master = "SELECT * FROM WORKER WHERE RFID_UID = (%s) AND STATUT = 'Formateur'"

					presence_master = 0

					presence_master = cur.execute(test_master, [RFID_UID])

					if presence_master == 1:

						# Create Excel (.xlsx) file -----------------------------------------------
						wb = Workbook()
						# select demo.xlsx
						sheet=wb.active

						print ("\nCreation du fichier...")
						sys.stdout.flush()
						lcd.lcd_clear()
						lcd.lcd_display_string(" Export  ", 1)
						lcd.lcd_display_string(" En Cours ", 2)
						sleep(1)



						result_sql = "SELECT * from LOG"
						cur.execute(result_sql)
						result_presence = cur.fetchall()


						# declare cursors for the rows of the .xlsx file
						# in openpyxl, you start counting from row 1 and row 1 contains the headers
						row_main_file = 2

						sheet.title = "Presence"

						sheet['A1'] = "Nom"
						sheet['B1'] = 'Prenom'
						sheet['C1'] = 'Groupe'
						sheet['D1'] = 'Entreprise'
						sheet['E1'] = 'Heure'


						for index, value in enumerate(result_presence):
							try:
								# extract from the tuple 'results' the value of each item and store it in every column for all rows
								sheet.cell(row=row_main_file, column=1).value = value[1]
								sheet.cell(row=row_main_file, column=2).value = value[2]
								sheet.cell(row=row_main_file, column=3).value = value[3]
								sheet.cell(row=row_main_file, column=4).value = value[4]
								sheet.cell(row=row_main_file, column=5).value = value[6]
							except:
								print("Error in line %s\n data=%s" % (index, value))
								row_main_file = row_main_file + 1

						workbook_name = "Presence du "
						wb.save('/home/pi/PointRFID/Export/'+workbook_name + DATE +".xlsx")

						try:
							sql_export = "UPDATE LOG SET EXPORT = 'True' WHERE EXPORT = 'False'"
							cur.execute(sql_export)
							db.commit()

							print ("\nMise A 1 de la valeur export...")
							sys.stdout.flush()

						except mysql.connector.Error as error :
							db.rollback()

							print("\nImpossible d\'executer la commande vers MySQL {}".format(error))
							sys.stdout.flush()
					else:
						print("\nCarte Non trouvee dans la base")
						sys.stdout.flush()
						lcd.lcd_clear()
						lcd.lcd_display_string(" Carte  ", 1)
						lcd.lcd_display_string(" Non Reconnu ", 2)
						sleep(1)

				except:
					print("\nCarte Non trouvee dans la base")
					sys.stdout.flush()

