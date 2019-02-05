from openpyxl import Workbook
import RPi.GPIO as GPIO    # Import Raspberry Pi GPIO library
from time import sleep     # Import the sleep function from the time module
import MFRC522
import signal
import sys
import os
import time
import mysql.connector
import MySQLdb
from mysql.connector import Error
from mysql.connector import errorcode
from datetime import datetime


GPIO.setwarnings(False)    # Ignore warning for now

GPIO.setmode(GPIO.BOARD)   # Use physical pin numbering
GPIO.setup(31, GPIO.OUT, initial=GPIO.LOW)   # Set pin 3 to be an output pin and set initial value to low (off) GREEN LED
GPIO.setup(33, GPIO.OUT, initial=GPIO.LOW)   # Set pin 5 to be an output pin and set initial value to low (off) ORANGE LED
GPIO.setup(35, GPIO.OUT, initial=GPIO.LOW)   # Set pin 7 to be an output pin and set initial value to low (off) RED LED

GPIO.setup(37, GPIO.OUT, initial=GPIO.LOW)




# Connect to DB -----------------------------------------------------------
db = MySQLdb.connect(host="localhost", user="root",passwd="Melec", db="RFID")
cur = db.cursor()


DATE = time.strftime("%Y-%m-%d %H:%M:%S")


continue_reading = True

# Fonction qui arrete la lecture proprement
def end_read(signal,frame):
	global continue_reading
	print ("\nLecture terminee")
	continue_reading = False
	GPIO.cleanup()

signal.signal(signal.SIGINT, end_read)
MIFAREReader = MFRC522.MFRC522()


while continue_reading:



	#os.system('pkill -9 -f ajout.py')

	# Detecter les tags
	(status,TagType) = MIFAREReader.MFRC522_Request(MIFAREReader.PICC_REQIDL)


	# Une carte est detectee
	if status == MIFAREReader.MI_OK:
		print ("\nCarte detectee")
		sys.stdout.flush()

	# Recuperation UID
	(status,uid) = MIFAREReader.MFRC522_Anticoll()

	if status == MIFAREReader.MI_OK:

		#Clignottement led verte

		uid2 = (str(uid[0])+str(uid[1])+str(uid[2])+str(uid[3]))

		print ("\nUID de la carte  :  "),uid2
		sys.stdout.flush()

		GPIO.output(31, GPIO.HIGH) # Turn on
		sleep(0.25)                  # Sleep for 1 second
		GPIO.output(31, GPIO.LOW)  # Turn off
		sleep(0.25)   
		GPIO.output(31, GPIO.HIGH) # Turn on
		sleep(0.25)                  # Sleep for 1 second
		GPIO.output(31, GPIO.LOW)  # Turn off
		sleep(0.25)   
		GPIO.output(31, GPIO.HIGH) # Turn on
		sleep(0.25)                  # Sleep for 1 second
		GPIO.output(31, GPIO.LOW)  # Turn off
		sleep(0.25)   
		
		try:
			print ("\nLecture de la base Eleve...")
			sys.stdout.flush()

			# Execute the SQL command
			recuperation = "SELECT * FROM ELEVE WHERE UID = (%s)"

			cur.execute(recuperation, [uid2])

			rows = cur.fetchall()

			ID = rows[0][0]
			NOM = rows[0][1]
			PRENOM = rows[0][2]
			EMAIL = rows[0][3]
			GROUPE = rows[0][4]
			ENTREPRISE = rows[0][5]
			UID = rows[0][6]
			HEURE = time.strftime("%Y-%m-%d %H:%M")

			presence = 0

			presence = "SELECT * FROM PRESENCE WHERE ID = (%s)"
			presence = cur.execute(presence, [ID])

			if presence == 1:
				#Suppression de la ligne
				try:

					print ("\nEleve deja present ")
					sys.stdout.flush()
					GPIO.output(31, GPIO.HIGH) # Turn on
					GPIO.output(37, GPIO.HIGH)
					sleep(0.25)
					GPIO.output(37, GPIO.LOW)
					sleep(0.25)
					GPIO.output(37, GPIO.HIGH)
					sleep(0.25)
					GPIO.output(37, GPIO.LOW)
					sleep(0.25)

				except:
					print ("\nErreur ")
					sys.stdout.flush()
					GPIO.output(33, GPIO.HIGH) # Turn on
					sleep(1)

			else:
				#Insertion de l'eleve
				try:
				
					insertion = """ INSERT INTO `PRESENCE`(`ID`, `NOM`, `PRENOM`, `GROUPE`, `ENTREPRISE`, `HEURE`) VALUES (%s,%s,%s,%s,%s,%s)"""

					insertion_var = (ID, NOM, PRENOM, GROUPE, ENTREPRISE, HEURE)
					
					result = cur.execute(insertion, insertion_var)

					data = cur.fetchall()

					db.commit()
					print ("\nEleve ajoutee a la base")
					sys.stdout.flush()

					GPIO.output(31, GPIO.HIGH) # Turn on
					GPIO.output(37, GPIO.HIGH)
					sleep(0.5)
					GPIO.output(37, GPIO.LOW)
					sleep(0.5)
					

				except mysql.connector.Error as error :
					db.rollback()

					print("\nImpossible d\'executer la commande vers MySQL {}".format(error))
					sys.stdout.flush()
					GPIO.output(35, GPIO.HIGH) # Turn on
					sleep(1)
		except:
			try:
				print ("\nLecture de la base Master...")
				sys.stdout.flush()
				# Execute the SQL command
				test_master = "SELECT * FROM MASTER WHERE UID = (%s)"

				presence_master = 0

				presence_master = cur.execute(test_master, [uid2])

				if presence_master == 1:

					GPIO.output(31, GPIO.HIGH) # Turn on
					GPIO.output(33, GPIO.HIGH) # Turn on
					sleep(1)


					# Create Excel (.xlsx) file -----------------------------------------------
					wb = Workbook()
					# select demo.xlsx
					sheet=wb.active

					print ("\nCreation du fichier...")
					sys.stdout.flush()


					result_sql = "SELECT * from PRESENCE"
					cur.execute(result_sql)
					result_presence = cur.fetchall()


					# declare cursors for the rows of the .xlsx file
					# in openpyxl, you start counting from row 1 and row 1 contains the headers
					row_main_file = 2

					sheet.title = "Presence"

					sheet['A1'] = "Nom"
					sheet['B1'] = 'Prenom'
					sheet['C1'] = 'Groupe'
					sheet['D1'] = 'Entreprise'
					sheet['E1'] = 'Heure'


					for index, value in enumerate(result_presence):
						try:
							# extract from the tuple 'results' the value of each item and store it in every column for all rows
							sheet.cell(row=row_main_file, column=1).value = value[1]
							sheet.cell(row=row_main_file, column=2).value = value[2]
							sheet.cell(row=row_main_file, column=3).value = value[3]
							sheet.cell(row=row_main_file, column=4).value = value[4]
							sheet.cell(row=row_main_file, column=5).value = value[5]
						except:
							print("Error in line %s\n data=%s" % (index, value))
						row_main_file = row_main_file + 1

					workbook_name = "Presence du "
					wb.save('/home/pi/RFID/'+workbook_name + DATE +".xlsx")
					# truncate tables and database ----------------------------------------------
					SQL = "TRUNCATE TABLE `PRESENCE`"
					cur.execute(SQL)
					db.commit()

					#Allumer Led Orange
					GPIO.output(35, GPIO.HIGH) # Turn on
					sleep(0.25)
					GPIO.output(35, GPIO.LOW) # Turn off
					sleep(0.25)
					GPIO.output(35, GPIO.HIGH) # Turn on
					sleep(0.25)
					GPIO.output(35, GPIO.LOW) # Turn off
					sleep(0.25)



			except:
				print("\nCarte Non trouvee dans la base")
				sys.stdout.flush()
				#Allumer Led Orange
				GPIO.output(33, GPIO.HIGH) # Turn on
				sleep(0.25)                  # Sleep for 1 second
				GPIO.output(33, GPIO.LOW)  # Turn off
				sleep(0.25)   
				GPIO.output(33, GPIO.HIGH) # Turn on
				sleep(0.25)                  # Sleep for 1 second
				GPIO.output(33, GPIO.LOW)  # Turn off
				sleep(0.25)  

		if status == MIFAREReader.MI_OK:
			#MIFAREReader.MFRC522_Read(8)
			GPIO.output(31, GPIO.LOW) # Turn off
			GPIO.output(33, GPIO.LOW) # Turn off
			GPIO.output(35, GPIO.LOW) # Turn off
			GPIO.output(37, GPIO.LOW) # Turn off
			print("\nTermine")
			MIFAREReader.MFRC522_StopCrypto1()

		else:
			print ("\nErreur de lecture de la carte RFID")
			sys.stdout.flush()
			#Clignottement Led Orange
			GPIO.output(35, GPIO.HIGH) # Turn on
			sleep(0.25)                  # Sleep for 1 second
			GPIO.output(35, GPIO.LOW)  # Turn off
			sleep(0.25)   
			GPIO.output(35, GPIO.HIGH) # Turn on
			sleep(0.25)                  # Sleep for 1 second
			GPIO.output(35, GPIO.LOW)  # Turn off
			sleep(0.25)  
